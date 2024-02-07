import os

from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# Set width and height the same
SD_W = 4/7
SD_H = 3


def main(img_path, xlsx_path):
    wb = openpyxl.Workbook()
    ws = wb.active

    img = Image.open(img_path)
    width, height = img.size

    for i in range(1, width + 1):
        ch = get_column_letter(i)
        ws.column_dimensions[ch].width = SD_W

    for j in range(1, height + 1):
        ws.row_dimensions[j].height = SD_H

    # Fill cell
    for i in range(1, width + 1):
        ch = get_column_letter(i)
        for j in range(1, height + 1):
            color = img.getpixel((i - 1, j - 1))
            if len(color) == 4:
                alpha = hex(color[3])[2:].zfill(2)
            else:
                alpha = 'FF'
            color_text = alpha + hex(color[0])[2:].zfill(2) + \
                hex(color[1])[2:].zfill(2) + hex(color[2])[2:].zfill(2)
            ws[f'{ch}{j}'].fill = PatternFill(
                start_color=color_text,
                end_color=color_text,
                fill_type='solid')

    wb.save(xlsx_path)


if __name__ == '__main__':
    img_path = r'example\test.jpg'
    name = os.path.splitext(os.path.basename(img_path))[0]
    xlsx_path = os.path.join('example', f'{name}.xlsx')
    main(img_path, xlsx_path)
